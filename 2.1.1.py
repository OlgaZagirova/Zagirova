import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

class Vacancy:
    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

    currency_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as File:
            reader = csv.reader(File)
            header = next(reader)
            header_length = len(header)
            for i in reader:
                if '' not in i and len(i) == header_length:
                    yield dict(zip(header, i))

    def get_statistic(self):
        salary = {}
        salary_of_vacancy = {}
        salary_city = {}
        count_vacancies = 0
        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_vacancies += 1
        number = dict([(key, len(value)) for key, value in salary.items()])
        number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy.items()])
        if not salary_of_vacancy:
            salary_of_vacancy = dict([(key, [0]) for key, value in salary.items()])
            number_by_name = dict([(key, 0) for key, value in number.items()])
        statistics = self.average(salary)
        statistics2 = self.average(salary_of_vacancy)
        statistics3 = self.average(salary_city)
        statistics4 = {}
        for i, j in salary_city.items():
            statistics4[i] = round(len(j) / count_vacancies, 4)
        statistics4 = list(filter(lambda x: x[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda x: x[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda x: x[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda x: x[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])

        return statistics, number, statistics2, number_by_name, statistics3, statistics5

    @staticmethod
    def average(dictionary):
        dict = {}
        for i, j in dictionary.items():
            dict[i] = int(sum(j) / len(j))
        return dict

    @staticmethod
    def increment(dict, key, amount):
        if key in dict:
            dict[key] += amount
        else:
            dict[key] = amount

    @staticmethod
    def print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics1))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))

class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        statistics1, statistics2, statistics3, statistics4, statistics5, statistics6 = dataset.get_statistic()
        dataset.print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        rep = Report(self.vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        rep.generate_excel()

class Report:
    def __init__(self, vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.statistics1 = statistics1
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def generate_excel(self):
        statistic = self.wb.active
        statistic.title = 'Статистика по годам'
        statistic.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.statistics1.keys():
            statistic.append([year, self.statistics1[year], self.statistics3[year], self.statistics2[year], self.statistics4[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column) > i:
                    if len(cell) > column[i]:
                        column[i] = len(cell)
                else:
                    column += [len(cell)]
        for i, j in enumerate(column, 1):
            statistic.column_dimensions[get_column_letter(i)].width = j + 2
        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            data.append([city1, value1, '', city2, value2])
        statistics = self.wb.create_sheet('Статистика по городам')
        for row in data:
            statistics.append(row)
        column = []
        for i in data:
            for x, y in enumerate(i):
                y = str(y)
                if len(column) > x:
                    if len(y) > column[x]:
                        column[x] = len(y)
                else:
                    column += [len(y)]
        for i, j in enumerate(column, 1):
            statistics.column_dimensions[get_column_letter(i)].width = j + 2
        font = Font(bold=True)
        for i in 'ABCDE':
            statistic[i + '1'].font = font
            statistics[i + '1'].font = font
        for i, _ in enumerate(self.statistics5):
            statistics['E' + str(i + 2)].number_format = '0.00%'
        side = Side(border_style='thin', color='00000000')
        for i in range(len(data)):
            for j in 'ABDE':
                statistics[j + str(i + 1)].border = Border(left=side, bottom=side, right=side, top=side)
        self.statistics1[1] = 1
        for i, _ in enumerate(self.statistics1):
            for j in 'ABCDE':
                statistic[j + str(i + 1)].border = Border(left=side, bottom=side, right=side, top=side)
        self.wb.save('report.xlsx')

InputConnect()