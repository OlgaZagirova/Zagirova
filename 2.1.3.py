import csv
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import pathlib
import pdfkit
from jinja2 import Environment, FileSystemLoader

class Vacancy:
    currency_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for i in reader:
                if '' not in i and len(i) == header_length:
                    yield dict(zip(header, i))

    def get_statistic(self):
        salary = {}
        salary_of_vacancy = {}
        salary_city = {}
        count_vacancies = 0
        for i in self.csv_reader():
            vacancy = Vacancy(i)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_vacancies += 1
        number = dict([(key, len(value)) for key, value in salary.items()])
        number_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy.items()])
        if not salary_of_vacancy:
            salary_of_vacancy = dict([(key, [0]) for key, value in salary.items()])
            number_by_name = dict([(key, 0) for key, value in number.items()])
        statistics1 = self.average(salary)
        statistics2 = self.average(salary_of_vacancy)
        statistics3 = self.average(salary_city)
        statistics4 = {}
        for i, j in salary_city.items():
            statistics4[i] = round(len(j) / count_vacancies, 4)
        statistics4 = list(filter(lambda x: x[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda x: x[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda x: x[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda x: x[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])

        return statistics1, number, statistics2, number_by_name, statistics3, statistics5

    @staticmethod
    def increment(dict, key, amount):
        if key in dict:
            dict[key] += amount
        else:
            dict[key] = amount

    @staticmethod
    def average(dictionary):
        dict = {}
        for i, j in dictionary.items():
            dict[i] = int(sum(j) / len(j))
        return dict

    @staticmethod
    def print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics1))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        dataset = DataSet(self.file_name, self.vacancy_name)
        statistics1, statistics2, statistics3, statistics4, statistics5, statistics6 = dataset.get_statistic()
        dataset.print_statistic(statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        rep = Report(self.vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6)
        rep.generate_excel()
        rep.generate_image()
        rep.generate_pdf()

class Report:
    def __init__(self, vacancy_name, statistics1, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.statistics1 = statistics1
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def generate_excel(self):
        statistic = self.wb.active
        statistic.title = 'Статистика по годам'
        statistic.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for x in self.statistics1.keys():
            statistic.append([x, self.statistics1[x], self.statistics3[x], self.statistics2[x], self.statistics4[x]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий', ' Количество вакансий - ' + self.vacancy_name]]
        column = []
        for x in data:
            for x, y in enumerate(x):
                if len(column) > x:
                    if len(y) > column[x]:
                        column[x] = len(y)
                else:
                    column += [len(y)]
        for i, j in enumerate(column, 1):
            statistic.column_dimensions[get_column_letter(i)].width = j + 2
        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            data.append([city1, value1, '', city2, value2])
        statistics = self.wb.create_sheet('Статистика по городам')
        for i in data:
            statistics.append(i)
        column = []
        for i in data:
            for x, y in enumerate(i):
                y = str(y)
                if len(column) > x:
                    if len(y) > column[x]:
                        column[x] = len(y)
                else:
                    column += [len(y)]
        for i, j in enumerate(column, 1):
            statistics.column_dimensions[get_column_letter(i)].width = j + 2
        font = Font(bold=True)
        for i in 'ABCDE':
            statistic[i + '1'].font = font
            statistics[i + '1'].font = font
        for i, _ in enumerate(self.statistics5):
            statistics['E' + str(i + 2)].number_format = '0.00%'
        side = Side(border_style='thin', color='00000000')
        for i in range(len(data)):
            for j in 'ABDE':
                statistics[j + str(i + 1)].border = Border(left=side, bottom=side, right=side, top=side)
        for i, _ in enumerate(self.statistics1):
            for j in 'ABCDE':
                statistic[j + str(i + 1)].border = Border(left=side, bottom=side, right=side, top=side)
        self.wb.save(filename='report.xlsx')

    def generate_image(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        bar1 = ax1.bar(np.array(list(self.statistics1.keys())) - 0.4, self.statistics1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.statistics1.keys())), self.statistics3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.statistics1.keys())) - 0.2, list(self.statistics1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)
        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.statistics2.keys())) - 0.4, self.statistics2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.statistics2.keys())), self.statistics4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()), prop={'size': 8})
        ax2.set_xticks(np.array(list(self.statistics2.keys())) - 0.2, list(self.statistics2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)
        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(i).replace(' ', '\n').replace('-', '-\n') for i in reversed(list(self.statistics5.keys()))]), list(reversed(list(self.statistics5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')
        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([x for x in self.statistics6.values()])
        ax4.pie(list(self.statistics6.values()) + [other], labels=list(self.statistics6.keys()) + ['Другие'], textprops={'fontsize': 6})
        plt.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        environment = Environment(loader=FileSystemLoader('.'))
        template = environment.get_template("index.html")
        stats = []
        for i in self.statistics1.keys():
            stats.append([i, self.statistics1[i], self.statistics2[i], self.statistics3[i], self.statistics4[i]])
        for i in self.statistics6:
            self.statistics6[i] = round(self.statistics6[i] * 100, 2)
        pdf_template = template.render({'name': self.vacancy_name, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                                        'stats': stats, 'stats5': self.statistics5, 'stats6': self.statistics6})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""}, configuration=config)

InputConnect()